# https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.read_sql_query.html#pandas.read_sql_query
# https://docs.microsoft.com/en-us/sql/machine-learning/data-exploration/python-dataframe-pandas?view=sql-server-ver15
# creator : RITA

# import all required packages
from mysql.connector import Error
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
import pandas as pd
import os
from pandas import ExcelWriter
from DB_Connection import MyDB

mydb = MyDB()

_file_name = 'Multiple_worksheets.xlsx'


def read_customer_data():
    try:
        name = []
        age = []
        cust_dic = {}
        conn = mydb.get_connection()
        sql_select_Query = "select * from Customer"
        cursor = conn.cursor()
        cursor.execute(sql_select_Query)
        records = cursor.fetchall()
        total = 0
        for row in records :
            name.append(row[0])
            age.append(row[1])
        cust_dic.update({'Name' : name, 'Age' : age})
        return cust_dic
    except Error as e:
        print("Error reading data from MySQL table", e)


def read_product_data():
    try:
        name = []
        price = []
        prod_dic = {}
        conn = mydb.get_connection()
        sql_select_Query = "select * from Product"
        cursor = conn.cursor()
        cursor.execute(sql_select_Query)
        records = cursor.fetchall()
        for row in records:
            name.append(row[0])
            price.append(row[1])
        prod_dic.update({'Name': name, 'Price':price})
        return prod_dic
    except Error as e:
        print("Error reading data from MySQL table", e)


def write_data(wb):
    _sheet_name1 = 'Customer_1'
    _sheet_name2 = 'Product_2'
    cust_data = read_customer_data()
    cust_df = pd.DataFrame(cust_data)
    prod_data = read_product_data()
    prod_df = pd.DataFrame(prod_data)
    with ExcelWriter(_file_name) as writer :
        cust_df.to_excel(writer, sheet_name= _sheet_name1)
        prod_df.to_excel(writer, sheet_name= _sheet_name2)


if os.path.exists(_file_name):
    wb = load_workbook(_file_name)
    write_data(wb)
else:
    print('new_workbook()')
    wb = Workbook()
    write_data(wb)




