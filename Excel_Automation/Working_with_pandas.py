# https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.read_sql_query.html#pandas.read_sql_query

# creator : RITA

# import all required packages
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
import pandas as pd
import os

month = datetime.today().month
year = datetime.today().year
_date = str(datetime.today()).split()[0]
dt = str((datetime.now() - timedelta(days=0)).strftime("%Y-%m-%d %H:%M:%S"))
_file_name = f'Efforts_{month}_{year}.xlsx'
_sheet_name = f'Data_{_date}'
_ticket_file = 'Data.xlsx'
_row = 0


# create New Workbook if Workbook doesn't exists
def new_workbook(_file_name):
    wb = Workbook()  # Workbook Object
    ws = wb.active  # Gets the active worksheet
    ws.title = _sheet_name  # Name the active worksheet

    # Writing the header columns
    ws['A1'] = 'Ticket No'
    ws['B1'] = 'Ref No'
    ws['C1'] = 'Corp ID'
    ws['D1'] = 'Activity Type'
    ws['E1'] = 'Activity'
    ws['F1'] = 'Timecard Entry Date'
    ws['G1'] = 'Effort'
    ws['H1'] = 'Complexity'
    ws['I1'] = 'AMorAD'
    ws['J1'] = 'SOW'
    ws['K1'] = 'Project'

    col_range = ws.max_column  # get max coulns in the worksheet

    # formatting the header columns, filling the color
    for col in range(1, col_range + 1):
        cell_header = ws.cell(1, col)
        cell_header.fill = PatternFill(start_color='e6f738', end_color='e6f738', fill_type="solid")

    wb.save(_file_name)  # save the workbook
    wb.close()  # close the workbook


def read_data_excel():
    # Checks if the file is present?
    if os.path.exists(_ticket_file):
        tickets = pd.read_excel(_ticket_file, usecols='A')  # Read the file with excel file
        tickets_list = ['', '0']  # List with some default values

        # Appending the cell value in the list
        for i in tickets['Ticket']:
            tickets_list.append(str(i).strip())

        return tickets_list
    else:
        wb = Workbook()  # Workbook object
        ws = wb.active  # Get Active Worksheet
        ws['A1'] = 'Ticket'  # Giving Header Name

        # Formatting the Header
        ws.cell(1, 1).fill = PatternFill(start_color='e6f738', end_color='e6f738', fill_type="solid")

        wb.save(_ticket_file)
        wb.close()


def create_list(_row):
    corp_id = 'user_name'
    project = 'project_name'
    complexity = 'Medium'
    activity_type = ''
    activity = ''
    reference = ''
    effort = 0
    AMorAD = ''
    SOW = ''
    my_tickets_record = []

    tickets = read_data_excel()  # Read Excel

    if tickets is None:
        print(f'No records in {_ticket_file}, fill in records and try again.')
    else:
        _rows = len(tickets) + 1

        for ticket in tickets:
            if ticket == '':
                activity_type = 'Meetings / Communication'
                activity = 'Mail Communication'
                effort = 1.5
            elif ticket == '0':
                activity_type = 'Service-Task'
                activity = 'DSTUM'
                effort = 1.5
            elif ticket[:3] == 'CHG':
                activity_type = 'Change Request'
                activity = 'Third party coordination'
                effort = 1
            else:
                activity_type = 'Incident'
                activity = 'Incident'
                effort = 0.75

            my_tickets_record.append([ticket, reference, corp_id, activity_type, activity, dt, effort, complexity, AMorAD, SOW, project])

        my_tickets_record.append([f'=SUM(G{_row}:G{_row + _rows})'])
        return my_tickets_record


def write_existing_wb(_file_name):
    existing_wb = load_workbook(_file_name)
    existing_ws = existing_wb.active
    _max_rows = existing_ws.max_row
    records = create_list(_max_rows)
    if records is not None:
        for row in records:
            existing_ws.append(row)
        existing_wb.save(_file_name)
        existing_wb.close()


if os.path.exists(_file_name):
    write_existing_wb(_file_name)
else:
    new_workbook(_file_name)
    write_existing_wb(_file_name)
